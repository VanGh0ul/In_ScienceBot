from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State,StatesGroup
from aiogram import types
from aiogram import types,Dispatcher
from create_bot import dp, bot
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove	
from openpyxl import Workbook
import openpyxl
import requests
import telebot
import json
import os

class FSMAdmin(StatesGroup):
	goal = State()
	name = State()
	card_view = State()

class FSMAdminGoal(StatesGroup):
	goal = State()

class FSMAdminDelete(StatesGroup):
	delete = State()


agentsName = ["",]
agentsID = ["",]

#Начало диалога



async def cm_start(message : types.Message):
	await FSMAdmin.goal.set()
	await message.reply('Введите цель поиска сотрудника:',reply_markup=types.ReplyKeyboardRemove())

#Команда старт
async def command_start(message : types.Message):

	try:
		keyboardStart = types.ReplyKeyboardMarkup(resize_keyboard=True)
		buttonbFindAgent = ["Найти сотрудника"]
		buttonsStart = ["Помощь", "Информация"]
		buttonsFile = ["Поиск по цели", "Удалить - цель по индексу"]
		keyboardStart.add(*buttonsStart).add(*buttonbFindAgent).add(*buttonsFile)
		await bot.send_message(message.from_user.id, 'Данный сервис предназначен для анализа и обработки сообщений💬 \nНажмите Найти сотрудника👨‍🔬 \nДалее введите Фамилию сотрудника для уточнения!', reply_markup=keyboardStart)

	except:
		await message.reply('❌Произошла ошибка,следите за информацией в :\nhttps://t.me/inscience_news')
		#await message.reply('Общение с ботом через ЛС, напишите ему:\nhttp://t.me/IN_SCIENCEBot')



#Просмотр списка
def get_number_of_elements(list):
	count = 0
	for element in list:
		count += 1
		if count > 5:
			break
	return count

def init_search_log(client_id):

	file_path = "goal_data\\" + str(client_id) + ".xlsx"

	wb = None
	if (os.path.exists(file_path)):

		wb = openpyxl.load_workbook(file_path)

	else:
		wb = Workbook()

	ws = wb.active
	ws["A1"] = "Пользователь искал"
	ws["B1"] = "Цель поиска"

	wb.save(file_path)


def append_client_info(client_id, column, val, is_family = False):

	file_path = "goal_data\\" + str(client_id) + ".xlsx"
	wb = openpyxl.load_workbook(file_path)

	ws = wb.active

	if (is_family):

		ws[column + str(len(ws[column]))] = val

	else:
		ws[column + str(len(ws[column]) + 1)] = val

	wb.save(file_path)

#Состояние 1
#Ответ от пользователя
#@dp.message_handler(state=FSMAdmin.name)

async def select_goal(message : types.Message, state: FSMContext):
	async with state.proxy() as data:
		
		keyboardBack = types.ReplyKeyboardMarkup(resize_keyboard=True)
		buttonsFind = ["Найти сотрудника"]
		buttonsBack = ["Меню"]
		keyboardBack.row(*buttonsFind).add(*buttonsBack)

		data['name'] = message.text
		if message.text.isdigit():
			await message.reply('❌Данные введены не корректно!', reply_markup=keyboardBack)
			await state.finish()
			return
		await FSMAdmin.next()
		await message.reply('Введите Фамилию сотрудника для поиска:')

		client_id = message.from_user.id

		init_search_log(client_id)
		append_client_info(client_id, "B", message.text)


async def select_agent(message : types.Message, state: FSMContext):
	async with state.proxy() as data:

		keyboardBack = types.ReplyKeyboardMarkup(resize_keyboard=True)
		buttonsFind = ["Найти сотрудника"]
		buttonsBack = ["Меню"]
		keyboardBack.row(*buttonsFind).add(*buttonsBack)

		data['name'] = message.text
		if message.text.isdigit():
			await message.reply('❌Данные введены не корректно!', reply_markup=keyboardBack)
			await state.finish()
			return

		url = "https://rinh-api.kovalev.team/employee/surname/"
		result = requests.get(url + data['name'])
		data = json.loads(result.content)

		parsName = "";
		parsID = "";
		countP = 1;
		item = 0

		#создание списков с ФИО и ид
		for item in data:
			
			parsName +=str(countP)+": - "+item["fullName"] + "\n"
			parsID = str(item["id"])
			agentsID.append(parsID)
			agentsName.append(parsName) 
			countP += 1
			if countP > 4:
				break
		#если всего 1 совпадение авто вывод карточки
		if (len(agentsID) == 2):

			append_client_info(message.from_user.id, "A", message.text, True)

			messageToInt = 1 
			employee_url = "https://rinh-api.kovalev.team/statistic/employee/findByEmployeeId?employeeId="
			result = requests.get(employee_url + agentsID[messageToInt])

			resultJson = json.loads(result.content)

			data = resultJson["employee"]
			wos = resultJson["dataTable"][0]
			scopus = resultJson["dataTable"][1]
			rinz = resultJson["dataTable"][2]

			articlesCount = wos["articlesCount"]

			parsID = "";
			countP = 0;

			avatarUrl = data["avatarUrl"]
			parsName = "Инициалы: " + str(data["fullName"])

			degree_url='https://rinh-api.kovalev.team/academic/degree/'
			title_url='https://rinh-api.kovalev.team/academic/title/'
			department_url='https://rinh-api.kovalev.team/department/'

			resultData =  {
				#"Id: ": data["id"],
				#"Ученое звание: ": data["academicTitleId"],
				#"Подразделение: ": data["positionId"],
				"Общий стаж: ": data["commonWorkExperience"],
				"Стаж по спец: ": data["specialtyWorkExperience"],
				"Кол-во статей в Wos: ": wos["articlesCount"],
				"Кол-во статей в Scopus: ": scopus["articlesCount"],
				"Кол-во статей в Elibrary: ": rinz["articlesCount"],
				"Email: ": data["email"],
				"Номер телефона: ": data["phone"]
			}

			resultDataDegree = {}
			resultDataTitle = {}
			resultDataDepartment = {}
		
			#проверка на пустоты
			if (data["positionId"] is not None):
				department = requests.get(department_url + str(data["positionId"]))

				print(department.content)

				resultJsonDepartment = json.loads(department.content)

				resultDataDepartment["Подразделение: "] = resultJsonDepartment["shortName"];

			if (data["academicTitleId"] is not None):
				title = requests.get(title_url + str(data["academicTitleId"]))

				print(title.content)

				resultJsonTitle = json.loads(title.content)

				resultDataTitle["Ученое звание: "] = resultJsonTitle["name"];

			if (data["academicDegreeId"] is not None):
				degree = requests.get(degree_url + str(data["academicDegreeId"]))

				print(degree.content)

				resultJsonDegree = json.loads(degree.content)

				resultDataDegree["Ученая степень: "] = resultJsonDegree["shortName"];


			if (data["academicDegree2Id"] is not None):
				degree = requests.get(degree_url + str(data["academicDegree2Id"]))

				print(degree.content)
				
				resultJsonDegree = json.loads(degree.content)

				resultDataDegree["Вторая ученная степень: "] = resultJsonDegree["shortName"];


			replyMessage = "";
			replyMessageDegree = "";
			replyMessageTitle = "";
			replyMessageDepartment = "";

			#сбор не пустой информации
			for key in resultData.keys():
				if (resultData[key] is not None):
					replyMessage += key + str(resultData[key]) + "\n"\

			for key in resultDataTitle.keys():
				if (resultDataTitle[key] is not None):
					replyMessageTitle += key + str(resultDataTitle[key]) + "\n"

			for key in resultDataDepartment.keys():
				if (resultDataDepartment[key] is not None):
					replyMessageDepartment += key + str(resultDataDepartment[key]) + "\n"

			for key in resultDataDegree.keys():
				if (resultDataDegree[key] is not None):
					replyMessageDegree += key + str(resultDataDegree[key]) + "\n"

			#инлайн кнопки
			inScienceUrl ='https://inscience.kovalev.team/institutes/user/'+str(data["id"])
			stankinUrl = str(data["authorUrlProfile"])
			scopusUrl ='https://www.scopus.com/authid/detail.uri?authorId='+str(data["scopusAuthorId"])
			wosUrl ='https://www.webofscience.com/wos/author/record/'+str(data["wosAuthorId"])
			rinzUrl = 'https://www.elibrary.ru/author_items.asp?authorid='+str(data["rinzAuthorId"])

			#Вывод карточки
			try:
				media = types.MediaGroup()	
				media.attach_photo(avatarUrl, " \n"+
					"\nКарточка сотрудника ⚛️💫"+
					"\n\n"+
					parsName+
					"\n\n"+
					replyMessageDegree+
					replyMessageDepartment+
					replyMessageTitle+
					"\n"+
					replyMessage
				)
				await bot.send_media_group(message.from_user.id, media=media)
				await bot.send_message(message.from_user.id,'✔️Нажмите "Найти сотрудника" для продолжения работы!', reply_markup=keyboardBack)

			#Вывод карточки без фото
			except:
				await message.reply("Карточка сотрудника ⚛💫"+
					"\n\n"+
					parsName+
					"\n\n"+
					replyMessageDegree+
					replyMessageDepartment+
					replyMessageTitle+
					"\n"+
					replyMessage,
					reply_markup=keyboardBack
				)

			#Связать id с инлайн кнопками
			inlineKb = types.InlineKeyboardMarkup(resize_keyboard=True)
			
			btn_inScienceUrl= types.InlineKeyboardButton(text='inScience', url=inScienceUrl)
			btn_stankinUrl= types.InlineKeyboardButton(text='Станкин', url=stankinUrl)
			btn_Wos = types.InlineKeyboardButton(text='Wos', url=wosUrl)
			btn_Scopus= types.InlineKeyboardButton(text='Scopus',url=scopusUrl)
			btn_Rinz= types.InlineKeyboardButton(text='Elibrary',  url=rinzUrl)

			inlineKb.row(btn_inScienceUrl).row(btn_stankinUrl).row(btn_Wos).row(btn_Scopus).row(btn_Rinz)

			await bot.send_message(message.chat.id, "Нажми на кнопку и перейди на сайты статей.", reply_markup = inlineKb)

			await bot.send_document(chat_id=message.chat.id, document=open("goal_data\\" + str(message.from_user.id) + ".xlsx", "rb"))

			await FSMAdmin.next()

			#очистить списки
			agentsName.clear()
			agentsID.clear()

			agentsID.append("",)
			agentsName.append("",) 

			await FSMAdmin.next()
		
		#не найдено не 1 сотрудника
		elif (len(agentsID) == 1):
			await message.reply("❌Не найдено ни одного сотрудника",reply_markup=keyboardBack)
			await state.finish()

		#если много сотрудников 2>
		elif (len(agentsID) > 2):
			
			listLenght = 5

			print("кнопки фиоо")
			await message.reply("⚠️Найдено несколько совпадений сотрудников \nПожалуйста выберите нужный вариант:")
			
			fieldBName = types.ReplyKeyboardMarkup(row_width=1)
			print(data)

			#Генератор кнопок с фио
			if (len(data) != 0):
				try:

					await bot.send_message(message.from_user.id, parsName)

					buttonsLength = listLenght
					
					if (len(data) < listLenght):
						buttonsLength = len(data)					

					fieldBName.add(*(types.KeyboardButton(data[i]["fullName"]) for i in range(buttonsLength)))
					await message.reply("✅Выберите сотрудника из списка для получения информации о нем!", reply_markup=fieldBName)

					agentsName.clear()
					agentsID.clear()

					agentsID.append("",)
					agentsName.append("",) 

					await FSMAdmin.next()

				except BadRequest:
					await message.reply("Авторов с такой фамилией найдено слишком много! Попробуй уточнить поиск", reply_markup=types.ReplyKeyboardRemove())

			else:
				await message.reply("Авторов с такой фамилией найдено не было! Попробуй еще раз", reply_markup=types.ReplyKeyboardRemove())
			#await bot.send_message(message.from_user.id,'✅Выберите сотрудника из списка для получения информации о нем!')

		else:
			try:
				await bot.send_message(message.from_user.id,parsName)
			except:
				await message.reply('❌Такого сотрудника нет,либо данные введены не корректно.')
				await state.finish()

		print('list id1:', agentsID," ")
		print('Updated list:', parsID," ")
		print("Number of elements in the list: ",get_number_of_elements(agentsID))

		#if (get_number_of_elements(agentsID))





#Принять одно фио из кнопки 
#Состояние 2
async def card_view(message: types.Message, state: FSMContext):

	#Вернуться назад (кнопки)
	keyboardBack = types.ReplyKeyboardMarkup(resize_keyboard=True)
	buttonsFind = ["Найти сотрудника"]
	buttonsBack = ["Меню"]

	keyboardBack.add(*buttonsFind).add(*buttonsBack)

	fullName = message.text

	append_client_info(message.from_user.id, "A", fullName, True)

	# fullname записывать

	url = "https://rinh-api.kovalev.team/employee/surname/"
	result = requests.get(url + fullName)

	employeeData = json.loads(result.content)

	#Проверка на пустоту или не вверные данные
	if (len(employeeData) == 0):
		await message.reply('❌Был произведен неправильный ввод!', reply_markup=keyboardBack)
		async with state.proxy() as data:
			await state.finish()
		return;
	

	employeeId = employeeData[0]["id"]

	employee_url = "https://rinh-api.kovalev.team/statistic/employee/findByEmployeeId?employeeId="
	result = requests.get(employee_url + str(employeeId))

	resultJson = json.loads(result.content)


	data = resultJson["employee"]
	wos = resultJson["dataTable"][0]
	scopus = resultJson["dataTable"][1]
	rinz = resultJson["dataTable"][2]
	# articlesData содержит массив articlesData.articlesCount - ошибка

	articlesCount = wos["articlesCount"]

	parsID = "";
	countP = 0;

	avatarUrl = data["avatarUrl"]
	parsName = "Инициалы: " + str(data["fullName"])

	#ссылки на списки degree title department
	degree_url='https://rinh-api.kovalev.team/academic/degree/'
	title_url='https://rinh-api.kovalev.team/academic/title/'
	department_url='https://rinh-api.kovalev.team/department/'

	resultData =  {
		#"Id: ": data["id"],
		#"Ученое звание: ": data["academicTitleId"],
		#"Подразделение: ": data["positionId"],
		"Общий стаж: ": data["commonWorkExperience"],
		"Стаж по спец: ": data["specialtyWorkExperience"],
		"Кол-во статей в Wos: ": wos["articlesCount"],
		"Кол-во статей в Scopus: ": scopus["articlesCount"],
		"Кол-во статей в Elibrary: ": rinz["articlesCount"],
		"Email: ": data["email"],
		"Номер телефона: ": data["phone"]
	}

	resultDataDegree = {}
	resultDataTitle = {}
	resultDataDepartment = {}

	#Проверка на пустоту

	if (data["positionId"] is not None):
		department = requests.get(department_url + str(data["positionId"]))

		print(department.content)

		resultJsonDepartment = json.loads(department.content)

		resultDataDepartment["Подразделение: "] = resultJsonDepartment["shortName"];

	if (data["academicTitleId"] is not None):
		title = requests.get(title_url + str(data["academicTitleId"]))

		print(title.content)

		resultJsonTitle = json.loads(title.content)

		resultDataTitle["Ученое звание: "] = resultJsonTitle["name"];

	if (data["academicDegreeId"] is not None):
		degree = requests.get(degree_url + str(data["academicDegreeId"]))

		print(degree.content)

		resultJsonDegree = json.loads(degree.content)

		resultDataDegree["Ученая степень: "] = resultJsonDegree["shortName"];


	if (data["academicDegree2Id"] is not None):
		degree = requests.get(degree_url + str(data["academicDegree2Id"]))

		print(degree.content)
		
		resultJsonDegree = json.loads(degree.content)

		resultDataDegree["Вторая ученная степень: "] = resultJsonDegree["shortName"];

	#Накопление не пустых данных
	replyMessage = "";
	replyMessageDegree = "";
	replyMessageTitle = "";
	replyMessageDepartment = "";

	for key in resultData.keys():
		if (resultData[key] is not None):
			replyMessage += key + str(resultData[key]) + "\n"\

	for key in resultDataTitle.keys():
		if (resultDataTitle[key] is not None):
			replyMessageTitle += key + str(resultDataTitle[key]) + "\n"

	for key in resultDataDepartment.keys():
		if (resultDataDepartment[key] is not None):
			replyMessageDepartment += key + str(resultDataDepartment[key]) + "\n"

	for key in resultDataDegree.keys():
		if (resultDataDegree[key] is not None):
			replyMessageDegree += key + str(resultDataDegree[key]) + "\n"
	#инлайн кнопки
	inScienceUrl ='https://inscience.kovalev.team/institutes/user/'+str(data["id"])
	stankinUrl = str(data["authorUrlProfile"])
	scopusUrl ='https://www.scopus.com/authid/detail.uri?authorId='+str(data["scopusAuthorId"])
	wosUrl ='https://www.webofscience.com/wos/author/record/'+str(data["wosAuthorId"])
	rinzUrl = 'https://www.elibrary.ru/author_items.asp?authorid='+str(data["rinzAuthorId"])

	#Вывод карточки+(фото)
	try:
		media = types.MediaGroup()	
		media.attach_photo(avatarUrl, " \n"+
			"\nКарточка сотрудника ⚛️💫"+
			"\n\n"+
			parsName+
			"\n\n"+
			replyMessageDegree+
			replyMessageDepartment+
			replyMessageTitle+
			"\n"+
			replyMessage
		)

		await bot.send_media_group(message.from_user.id, media=media)
		await bot.send_message(message.from_user.id,'✔️Нажмите "Найти сотрудника" для продолжения работы!', reply_markup=keyboardBack)
	except:
		#Вывод карточки(без фото)
		await message.reply("Карточка сотрудника ⚛💫"+
			"\n\n"+
			parsName+
			"\n\n"+
			replyMessageDegree+
			replyMessageDepartment+
			replyMessageTitle+
			"\n"+
			replyMessage,
			reply_markup=keyboardBack
		)

	#инлайн кнопки
	inlineKb = types.InlineKeyboardMarkup(resize_keyboard=True)
	
	btn_inScienceUrl= types.InlineKeyboardButton(text='inScience', url=inScienceUrl)
	btn_stankinUrl= types.InlineKeyboardButton(text='Станкин', url=stankinUrl)
	btn_Wos = types.InlineKeyboardButton(text='Wos', url=wosUrl)
	btn_Scopus= types.InlineKeyboardButton(text='Scopus',url=scopusUrl)
	btn_Rinz= types.InlineKeyboardButton(text='Elibrary',  url=rinzUrl)

	inlineKb.row(btn_inScienceUrl).row(btn_stankinUrl).row(btn_Wos).row(btn_Scopus).row(btn_Rinz)

	await bot.send_message(message.chat.id, "Нажми на кнопку что бы перейти на сайты статей.", reply_markup = inlineKb)

	await bot.send_document(chat_id=message.chat.id, document=open("goal_data\\" + str(message.from_user.id) + ".xlsx", "rb"))

	await FSMAdmin.next()
	#Закончить машину состояний
	async with state.proxy() as data:
		await state.finish()

async def cmGoal_start(message : types.Message):
	await FSMAdminGoal.goal.set()
	await message.reply('Введите по какой цели делать поиск:',reply_markup=types.ReplyKeyboardRemove())
	
async def goal_Find(message : types.Message, state: FSMContext):
	async with state.proxy() as data:
		
		keyboardBack = types.ReplyKeyboardMarkup(resize_keyboard=True)
		buttonsBack = ["Меню"]
		keyboardBack.add(*buttonsBack)

		data['name'] = message.text
		if message.text.isdigit():
			await message.reply('❌Данные введены не корректно!', reply_markup=keyboardBack)
			await state.finish()
			return
		print(message.text)

		file_path = "goal_data\\" + str(message.from_user.id) + ".xlsx"

		if (os.path.exists(file_path) is not True):
			await bot.send_message(message.from_user.id,"Файл списка запросов не найден", reply_markup=keyboardBack)
			await state.finish()
			return

		wb = openpyxl.load_workbook(file_path, read_only=True)
		ws = wb.active

		result_is_empty = True
		result = "Были найдены следующие запросы:\n"

		for row in ws.iter_rows(2):
			for cell in row:
				if cell.value == message.text:

					found = ws.cell(row=cell.row, column=1).value

					if (found is not None):
						result_is_empty = False
						result += found + "\n";

		wb.close()

		if (result_is_empty):
			result = "Запросы не были найдены"

		await bot.send_message(message.from_user.id, result, reply_markup=keyboardBack)

		await FSMAdmin.next()

	await FSMAdminGoal.next()
	#Закончить машину состояний
	async with state.proxy() as data:
		await state.finish()

async def cmDelete_start(message : types.Message):
	await FSMAdminDelete.delete.set()
	await message.reply('Введите № строки которую нужно удалить:',reply_markup=types.ReplyKeyboardRemove())
	
async def delete_Find(message : types.Message, state: FSMContext):
	async with state.proxy() as data:
		
		keyboardBack = types.ReplyKeyboardMarkup(resize_keyboard=True)
		buttonsBack = ["Меню"]
		keyboardBack.add(*buttonsBack)

		data['name'] = message.text

		file_path = "goal_data\\" + str(message.from_user.id) + ".xlsx"

		if (os.path.exists(file_path) is not True):
			await bot.send_message(message.from_user.id,"Файл списка запросов не найден", reply_markup=keyboardBack)
			await state.finish()
			return

		wb = openpyxl.load_workbook(file_path)

		try: 

			ws = wb.active

			row_ind = None

			try:

				row_ind = int(message.text)

			except Exception as ex:
				print(ex)
				await bot.send_message(message.from_user.id,"Значение-"+ message.text+"-является не подходящим для удаление", reply_markup=keyboardBack)
				await state.finish()
				return

			if (row_ind == 1):
				await bot.send_message(message.from_user.id,"Введите индекс начиная с 2", reply_markup=keyboardBack)
				await state.finish()
				return

			ws.delete_rows(row_ind)

			wb.save(file_path)

			await bot.send_message(message.from_user.id,"✅Удаление прошло упешно, просмотрите файл", reply_markup=keyboardBack)
			await bot.send_document(chat_id=message.chat.id, document=open("goal_data\\" + str(message.from_user.id) + ".xlsx", "rb"))

		finally:
			wb.close();

	await FSMAdminDelete.next()
	#Закончить машину состояний
	async with state.proxy() as data:
		await state.finish()

#Регистр хендов
def register_handlers_agentsInfo(dp : Dispatcher):
	dp.register_message_handler(cm_start, lambda message: message.text == "Найти сотрудника")
	dp.register_message_handler(cmGoal_start, lambda message: message.text == "Поиск по цели")
	dp.register_message_handler(cmDelete_start, lambda message: message.text == "Удалить - цель по индексу")
	dp.register_message_handler(command_start, lambda message: message.text == "Меню")
	dp.register_message_handler(delete_Find, state=FSMAdminDelete.delete)
	dp.register_message_handler(goal_Find, state=FSMAdminGoal.goal)
	dp.register_message_handler(select_goal, state=FSMAdmin.goal)
	dp.register_message_handler(select_agent, state=FSMAdmin.name)
	dp.register_message_handler(card_view, state=FSMAdmin.card_view)
	# dp.register_message_handler(select_agentNum, state=FSMAdmin.num)
	# dp.register_message_handler()
#Ответ от пользователя
